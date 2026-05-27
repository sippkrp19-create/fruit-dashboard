from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


SRC = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_สรุปมีรายรับ.xlsx")
OUT = SRC
LIST_SHEET = "รายการดรอปดาวน์"

GREEN = "1F4E3D"
GREEN_LIGHT = "E9F4EC"
GOLD_LIGHT = "FFF3D6"
WHITE = "FFFFFF"
GRAY_TEXT = "60706A"

GRADE_OPTIONS = [
    "เกรด1",
    "เกรด2",
    "เกรด3",
    "เกรด4",
    "เกรดพุ่ม",
    "เกรดปุก",
    "เกรดดำ",
    "เกรดร่วง",
]

YEAR_OPTIONS = list(range(2569, 2581))
MONTH_OPTIONS = [
    "ทั้งหมด",
    "ม.ค.",
    "ก.พ.",
    "มี.ค.",
    "เม.ย.",
    "พ.ค.",
    "มิ.ย.",
    "ก.ค.",
    "ส.ค.",
    "ก.ย.",
    "ต.ค.",
    "พ.ย.",
    "ธ.ค.",
]


def values_from_range(ws, col: str, start_row: int = 2, max_row: int = 200) -> list:
    values = []
    for row in range(start_row, max_row + 1):
        value = ws[f"{col}{row}"].value
        if value in (None, ""):
            continue
        if value not in values:
            values.append(value)
    return values


def merge_unique(*groups) -> list:
    out = []
    for group in groups:
        for value in group:
            if value in (None, ""):
                continue
            if value not in out:
                out.append(value)
    return out


def reset_validation(ws, target_cell: str, formula: str):
    for dv in list(ws.data_validations.dataValidation):
        if target_cell in dv.cells:
            dv.ranges.remove(target_cell)

    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.showErrorMessage = False
    dv.errorTitle = "ค่านี้ไม่มีในรายการ"
    dv.error = "ถ้าต้องการเพิ่มตัวเลือก ให้ไปเพิ่มในชีต รายการดรอปดาวน์"
    ws.add_data_validation(dv)
    dv.add(ws[target_cell])


def write_list_sheet(wb):
    form = wb["เพิ่มข้อมูล"]
    if LIST_SHEET in wb.sheetnames:
        ws = wb[LIST_SHEET]
        ws.delete_rows(1, max(ws.max_row, 1))
    else:
        ws = wb.create_sheet(LIST_SHEET, 1)

    lists = {
        "ปี/ฤดูกาล": YEAR_OPTIONS,
        "สินค้า": values_from_range(form, "AA"),
        "ประเภทรายการ": values_from_range(form, "AB"),
        "สายพันธุ์/เกรด": merge_unique(GRADE_OPTIONS, values_from_range(form, "AC")),
        "ชื่อ/รายการ": values_from_range(form, "AD"),
        "วิธีรับ/จ่ายเงิน": values_from_range(form, "AE"),
        "สถานะรายการ": values_from_range(form, "AF"),
        "เดือน": MONTH_OPTIONS,
    }

    ws.sheet_view.showGridLines = False
    ws["A1"] = "รายการดรอปดาวน์"
    ws["A1"].font = Font(size=18, bold=True, color=GREEN)
    ws["A2"] = "เพิ่ม/ลบตัวเลือกได้ในคอลัมน์ด้านล่าง แล้วกลับไปใช้ dropdown ในชีต เพิ่มข้อมูล"
    ws["A2"].font = Font(italic=True, color=GRAY_TEXT)

    for col_idx, (header, values) in enumerate(lists.items(), 1):
        cell = ws.cell(4, col_idx, header)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center")
        for row_idx, value in enumerate(values, 5):
            ws.cell(row_idx, col_idx, value)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max(18, min(34, len(header) + 12))

    for row in range(5, 105):
        for col in range(1, len(lists) + 1):
            ws.cell(row, col).fill = PatternFill("solid", fgColor=GOLD_LIGHT)

    return ws


def update_workbook():
    wb = openpyxl.load_workbook(SRC)
    list_ws = write_list_sheet(wb)
    form = wb["เพิ่มข้อมูล"]

    # Use broad list ranges so the user can add future values without changing validation rules.
    reset_validation(form, "C5", f"'{LIST_SHEET}'!$A$5:$A$104")
    reset_validation(form, "C6", f"'{LIST_SHEET}'!$B$5:$B$104")
    reset_validation(form, "C7", f"'{LIST_SHEET}'!$C$5:$C$104")
    reset_validation(form, "C8", f"'{LIST_SHEET}'!$D$5:$D$104")
    reset_validation(form, "C9", f"'{LIST_SHEET}'!$E$5:$E$104")
    reset_validation(form, "F6", f"'{LIST_SHEET}'!$F$5:$F$104")
    reset_validation(form, "F7", f"'{LIST_SHEET}'!$G$5:$G$104")

    if "สรุป" in wb.sheetnames:
        summary = wb["สรุป"]
        reset_validation(summary, "B3", f"'{LIST_SHEET}'!$A$5:$A$104")
        reset_validation(summary, "D3", f"'{LIST_SHEET}'!$H$5:$H$104")

    if form["C5"].value not in YEAR_OPTIONS:
        form["C5"] = 2569

    list_ws.freeze_panes = "A5"
    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print(update_workbook())
