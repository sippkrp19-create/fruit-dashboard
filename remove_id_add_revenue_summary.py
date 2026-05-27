from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


src = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_วางค่าได้.xlsx")
out = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_ไม่มีรหัส.xlsx")

wb = openpyxl.load_workbook(src)

db = wb["ฐานข้อมูล"]
form = wb["เพิ่มข้อมูล"]
summary = wb["สรุป"]

green = "1F4E3D"
white = "FFFFFF"
light_green = "E9F4EC"

# Remove transaction ID column from database.
if db.cell(4, 1).value == "รหัสรายการ":
    db.delete_cols(1)

# Refresh database formulas after column shift.
# New columns: A วันที่, B รอบ, C สินค้า, D ประเภท, E สายพันธุ์, F ชื่อ,
# G ราคา, H ปริมาณ, I รายรับ, J รายจ่าย, K ค่าตัด, L เบิก, M สุทธิ,
# N วิธี, O สถานะ, P วันที่บันทึก, Q หมายเหตุ, R แหล่งที่มา
for row in range(5, db.max_row + 1):
    db.cell(row, 9, f'=IF($D{row}="รายรับ",$G{row}*$H{row},0)')
    db.cell(row, 10, f'=IF($D{row}="รายจ่าย",$G{row}*$H{row},0)')
    db.cell(row, 13, f"=I{row}-J{row}-K{row}")
    db.cell(row, 1).number_format = "d/m/yyyy"
    db.cell(row, 16).number_format = "d/m/yyyy"

db.tables.clear()
db_ref = f"A4:R{db.max_row}"
table = Table(displayName="tblDatabase", ref=db_ref)
table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True, showColumnStripes=False)
db.add_table(table)
db.auto_filter.ref = db_ref
db.freeze_panes = "A5"

# Rebuild form preview row without ID.
headers = [
    "วันที่", "รอบ/ฤดูกาล", "สินค้า", "ประเภทรายการ", "สายพันธุ์/เกรด",
    "ชื่อ/รายการ", "ราคาต่อกก.", "ปริมาณกก.", "รายรับ", "รายจ่าย",
    "ค่าตัด", "เบิก", "สุทธิ", "วิธีรับ/จ่ายเงิน", "สถานะรายการ",
    "วันที่บันทึก", "หมายเหตุ", "แหล่งที่มา",
]
for col in range(1, 20):
    form.cell(17, col).value = None
    form.cell(18, col).value = None
for col, header in enumerate(headers, 1):
    cell = form.cell(17, col, header)
    cell.fill = PatternFill("solid", fgColor=green)
    cell.font = Font(bold=True, color=white)
    cell.alignment = Alignment(horizontal="center")

preview = [
    '=TEXT(\'เพิ่มข้อมูล\'!$C$4,"d/m/yyyy")',
    "='เพิ่มข้อมูล'!$C$5",
    "='เพิ่มข้อมูล'!$C$6",
    "='เพิ่มข้อมูล'!$C$7",
    '=IF(\'เพิ่มข้อมูล\'!$C$8<>"",\'เพิ่มข้อมูล\'!$C$8,"")',
    '=IF(\'เพิ่มข้อมูล\'!$C$9<>"",\'เพิ่มข้อมูล\'!$C$9,"")',
    "='เพิ่มข้อมูล'!$C$10",
    "='เพิ่มข้อมูล'!$C$11",
    "='เพิ่มข้อมูล'!$F$10",
    "='เพิ่มข้อมูล'!$F$11",
    "='เพิ่มข้อมูล'!$F$4",
    "='เพิ่มข้อมูล'!$F$5",
    "='เพิ่มข้อมูล'!$F$12",
    "='เพิ่มข้อมูล'!$F$6",
    "='เพิ่มข้อมูล'!$F$7",
    '=TEXT(TODAY(),"d/m/yyyy")',
    "='เพิ่มข้อมูล'!$F$8",
    '"เพิ่มข้อมูล"',
]
for col, formula in enumerate(preview, 1):
    cell = form.cell(18, col, formula)
    cell.fill = PatternFill("solid", fgColor=light_green)
for col in [1, 16]:
    form.cell(18, col).number_format = "@"
for col in [7, 8, 9, 10, 11, 12, 13]:
    form.cell(18, col).number_format = "#,##0.00"
form["A20"] = "วิธีใช้: กรอกข้อมูลให้ครบ แล้วคัดลอกแถว 18 ไปวางท้ายฐานข้อมูลด้วย Paste Values / วางค่าเท่านั้น"

# Rebuild Summary formulas to match shifted database columns and include revenue in product summary.
MAX_ROW = 2000
def filtered_sum(col, extra=""):
    base = (
        f"'ฐานข้อมูล'!${col}$5:${col}${MAX_ROW},"
        f"'ฐานข้อมูล'!$O$5:$O${MAX_ROW},\"ปกติ\","
        f"'ฐานข้อมูล'!$B$5:$B${MAX_ROW},$B$3"
    )
    if extra:
        base += f",{extra}"
    start = "DATE($B$3-543,$E$3,1)"
    month = (
        f"'ฐานข้อมูล'!${col}$5:${col}${MAX_ROW},"
        f"'ฐานข้อมูล'!$O$5:$O${MAX_ROW},\"ปกติ\","
        f"'ฐานข้อมูล'!$B$5:$B${MAX_ROW},$B$3"
    )
    if extra:
        month += f",{extra}"
    month += f",'ฐานข้อมูล'!$A$5:$A${MAX_ROW},\">=\"&{start},'ฐานข้อมูล'!$A$5:$A${MAX_ROW},\"<\"&EDATE({start},1)"
    return f'=IF($E$3="ทั้งหมด",SUMIFS({base}),SUMIFS({month}))'

summary.tables.clear()
summary["A4"] = "KPI"
summary["B4"] = "ค่า"
summary["A5"] = "รายรับรวม"
summary["B5"] = filtered_sum("I")
summary["A6"] = "รายจ่ายรวม"
summary["B6"] = f"={filtered_sum('J')[1:]}+{filtered_sum('K')[1:]}"
summary["A7"] = "กำไรสุทธิ"
summary["B7"] = "=B5-B6"
summary["A8"] = "ปริมาณกก.รวม"
summary["B8"] = filtered_sum("H")
summary["A9"] = "ราคาเฉลี่ย/กก."
summary["B9"] = "=IFERROR(B5/B8,0)"
summary["A10"] = "ค่าตัดรวม"
summary["B10"] = filtered_sum("K")
summary["A11"] = "เบิกรวม"
summary["B11"] = filtered_sum("L")

for col, header in enumerate(["สินค้า", "รายรับ", "รายจ่าย", "สุทธิ"], 4):
    summary.cell(4, col, header)
products = ["ทุเรียน", "มังคุด", "ลองกอง", "ยังไม่ชัว"]
for idx, product in enumerate(products, 5):
    summary.cell(idx, 4, product)
    prod = f"'ฐานข้อมูล'!$C$5:$C${MAX_ROW},D{idx}"
    summary.cell(idx, 5, filtered_sum("I", prod))
    summary.cell(idx, 6, f"={filtered_sum('J', prod)[1:]}+{filtered_sum('K', prod)[1:]}")
    summary.cell(idx, 7, f"=E{idx}-F{idx}")

# Expense by name after column shift.
for r in range(5, 18):
    if summary.cell(r, 9).value:
        name_crit = f"'ฐานข้อมูล'!$F$5:$F${MAX_ROW},I{r}"
        summary.cell(r, 10, filtered_sum("J", name_crit))

for month_num, row in enumerate(range(16, 28), 1):
    start = f"DATE($B$3-543,{month_num},1)"
    next_m = f"EDATE({start},1)"
    summary.cell(row, 1, f'=TEXT({start},"mmm yyyy")')
    summary.cell(row, 2, f'=SUMIFS(\'ฐานข้อมูล\'!$I$5:$I${MAX_ROW},\'ฐานข้อมูล\'!$O$5:$O${MAX_ROW},"ปกติ",\'ฐานข้อมูล\'!$B$5:$B${MAX_ROW},$B$3,\'ฐานข้อมูล\'!$A$5:$A${MAX_ROW},">="&{start},\'ฐานข้อมูล\'!$A$5:$A${MAX_ROW},"<"&{next_m})')
    summary.cell(row, 3, f'=SUMIFS(\'ฐานข้อมูล\'!$J$5:$J${MAX_ROW},\'ฐานข้อมูล\'!$O$5:$O${MAX_ROW},"ปกติ",\'ฐานข้อมูล\'!$B$5:$B${MAX_ROW},$B$3,\'ฐานข้อมูล\'!$A$5:$A${MAX_ROW},">="&{start},\'ฐานข้อมูล\'!$A$5:$A${MAX_ROW},"<"&{next_m})+SUMIFS(\'ฐานข้อมูล\'!$K$5:$K${MAX_ROW},\'ฐานข้อมูล\'!$O$5:$O${MAX_ROW},"ปกติ",\'ฐานข้อมูล\'!$B$5:$B${MAX_ROW},$B$3,\'ฐานข้อมูล\'!$A$5:$A${MAX_ROW},">="&{start},\'ฐานข้อมูล\'!$A$5:$A${MAX_ROW},"<"&{next_m})')
    summary.cell(row, 4, f"=B{row}-C{row}")

for row in [4, 15]:
    for col in range(1, 11):
        cell = summary.cell(row, col)
        if cell.value:
            cell.fill = PatternFill("solid", fgColor=green)
            cell.font = Font(bold=True, color=white)
            cell.alignment = Alignment(horizontal="center")

for ref, name in [("A4:B11", "tblKPI"), ("D4:G8", "tblByProduct"), ("I4:J17", "tblExpenseByName"), ("A15:D27", "tblByMonth")]:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True, showColumnStripes=False)
    summary.add_table(table)

for sheet in wb.worksheets:
    sheet.protection.sheet = False
    sheet.protection.enable = False

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(out)
print(out)
