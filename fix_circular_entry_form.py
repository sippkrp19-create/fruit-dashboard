from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill


src = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_ฟอร์มตรง.xlsx")
out = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_ไม่วนสูตร.xlsx")

wb = openpyxl.load_workbook(src)
ws = wb["เพิ่มข้อมูล"]
db = wb["ฐานข้อมูล"]

# Preview row: use explicit sheet-qualified references so a normal paste into
# ฐานข้อมูล will not become self-referential/circular. For permanent history,
# users should still paste values.
formulas = [
    '=TEXT(COUNTA(\'ฐานข้อมูล\'!$A:$A)-3+1,"TXN-0000")',
    "='เพิ่มข้อมูล'!$C$4",
    "='เพิ่มข้อมูล'!$C$5",
    "='เพิ่มข้อมูล'!$C$6",
    "='เพิ่มข้อมูล'!$C$7",
    '=IF(\'เพิ่มข้อมูล\'!$C$8<>"",\'เพิ่มข้อมูล\'!$C$8,"")',
    '=IF(\'เพิ่มข้อมูล\'!$C$9<>"",\'เพิ่มข้อมูล\'!$C$9,"")',
    "='เพิ่มข้อมูล'!$C$10",
    "='เพิ่มข้อมูล'!$C$11",
    "='เพิ่มข้อมูล'!$F$10",
    "='เพิ่มข้อมูล'!$F$11",
    "='เพิ่มข้อมูล'!$F$4",
    "='เพิ่มข้อมูล'!$F$5",
    "='เพิ่มข้อมูล'!$F$12",
    "='เพิ่มข้อมูล'!$F$6",
    "='เพิ่มข้อมูล'!$F$7",
    "=TODAY()",
    "='เพิ่มข้อมูล'!$F$8",
    '"เพิ่มข้อมูล"',
]
for col, formula in enumerate(formulas, 1):
    ws.cell(18, col, formula)

ws["A20"] = "วิธีใช้ที่ปลอดภัย: กรอกข้อมูลให้พร้อม แล้วคัดลอกแถว 18 ไปวางท้ายฐานข้อมูลด้วย Paste Values / วางค่าเท่านั้น"
ws["A20"].font = Font(italic=True, bold=True, color="8A3B35")
ws["A21"] = "ถ้าวางแบบปกติ สูตรจะยังลิงก์กับฟอร์มอยู่; ถ้าต้องการเก็บประวัติถาวรให้วางเป็นค่าเท่านั้น"
ws["A21"].font = Font(italic=True, color="60706A")
try:
    ws.unmerge_cells("A20:S20")
except Exception:
    pass
try:
    ws.unmerge_cells("A21:S21")
except Exception:
    pass
ws.merge_cells("A20:S20")
ws.merge_cells("A21:S21")

# Repair any accidental pasted preview formulas already present in the database:
# keep only row-local calculation formulas in J, K, N; other formula columns are
# converted to blanks so they cannot point back to the form or themselves.
for row in range(5, db.max_row + 1):
    for col in range(1, 20):
        cell = db.cell(row, col)
        if isinstance(cell.value, str) and cell.value.startswith("=") and col not in [10, 11, 14]:
            cell.value = None
    db.cell(row, 10, f'=IF($E{row}="รายรับ",$H{row}*$I{row},0)')
    db.cell(row, 11, f'=IF($E{row}="รายจ่าย",$H{row}*$I{row},0)')
    db.cell(row, 14, f"=J{row}-K{row}-L{row}")

for cell in ["B18", "Q18"]:
    ws[cell].number_format = "d/m/yyyy"
for col in [8, 9, 10, 11, 12, 13, 14]:
    ws.cell(18, col).number_format = "#,##0.00"

for sheet in wb.worksheets:
    sheet.protection.sheet = False
    sheet.protection.enable = False

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(out)
print(out)
