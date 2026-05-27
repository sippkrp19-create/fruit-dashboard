from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


src = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_ไม่มีรหัส.xlsx")
out = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_สรุปมีรายรับ.xlsx")

wb = openpyxl.load_workbook(src)
ws = wb["สรุป"]

green = "1F4E3D"
white = "FFFFFF"
light_green = "E9F4EC"
MAX_ROW = 2000

# Make sure no nearby column is hidden or too narrow.
for col, width in {"D": 16, "E": 16, "F": 16, "G": 16}.items():
    ws.column_dimensions[col].hidden = False
    ws.column_dimensions[col].width = width

# Remove and recreate only the product summary table if it exists.
if "tblByProduct" in ws.tables:
    del ws.tables["tblByProduct"]

headers = ["สินค้า", "รายรับ", "รายจ่าย", "สุทธิ"]
for col, header in enumerate(headers, 4):
    cell = ws.cell(4, col, header)
    cell.fill = PatternFill("solid", fgColor=green)
    cell.font = Font(bold=True, color=white)
    cell.alignment = Alignment(horizontal="center", vertical="center")

products = ["ทุเรียน", "มังคุด", "ลองกอง", "ยังไม่ชัว"]


def filtered_sum(col, extra=""):
    base = (
        f"'ฐานข้อมูล'!${col}$5:${col}${MAX_ROW},"
        f"'ฐานข้อมูล'!$O$5:$O${MAX_ROW},\"ปกติ\","
        f"'ฐานข้อมูล'!$B$5:$B${MAX_ROW},$B$3"
    )
    if extra:
        base += f",{extra}"
    start = "DATE($B$3-543,$E$3,1)"
    month = (
        f"'ฐานข้อมูล'!${col}$5:${col}${MAX_ROW},"
        f"'ฐานข้อมูล'!$O$5:$O${MAX_ROW},\"ปกติ\","
        f"'ฐานข้อมูล'!$B$5:$B${MAX_ROW},$B$3"
    )
    if extra:
        month += f",{extra}"
    month += f",'ฐานข้อมูล'!$A$5:$A${MAX_ROW},\">=\"&{start},'ฐานข้อมูล'!$A$5:$A${MAX_ROW},\"<\"&EDATE({start},1)"
    return f'=IF($E$3="ทั้งหมด",SUMIFS({base}),SUMIFS({month}))'


for idx, product in enumerate(products, 5):
    ws.cell(idx, 4, product)
    prod = f"'ฐานข้อมูล'!$C$5:$C${MAX_ROW},D{idx}"
    ws.cell(idx, 5, filtered_sum("I", prod))  # รายรับ
    ws.cell(idx, 6, f"={filtered_sum('J', prod)[1:]}+{filtered_sum('K', prod)[1:]}")  # รายจ่าย + ค่าตัด
    ws.cell(idx, 7, f"=E{idx}-F{idx}")  # สุทธิ
    for col in range(4, 8):
        cell = ws.cell(idx, col)
        if idx % 2 == 0:
            cell.fill = PatternFill("solid", fgColor=light_green)
        if col >= 5:
            cell.number_format = "#,##0.00"

table = Table(displayName="tblByProduct", ref="D4:G8")
table.tableStyleInfo = TableStyleInfo(
    name="TableStyleMedium4",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False,
)
ws.add_table(table)

for sheet in wb.worksheets:
    sheet.protection.sheet = False
    sheet.protection.enable = False

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(out)
print(out)
