from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


src = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชีและแดชบอร์ด_แก้ไขได้.xlsx")
out = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชีและแดชบอร์ด_เลือกเดือนปี.xlsx")

wb = openpyxl.load_workbook(src)
dash = wb["Dashboard"]
summary = wb["สรุป"]

green_dark = "1F4E3D"
gold_light = "FFF3D6"
green_light = "E9F4EC"
white = "FFFFFF"

# Clean and improve the filter controls.
dash["A3"] = "ปี"
dash["B3"] = 2569
dash["C3"] = "เดือน"
dash["D3"] = "ทั้งหมด"
dash["E3"] = '=IF(D3="ทั้งหมด","ทั้งหมด",MATCH(D3,$O$1:$O$12,0))'
dash.column_dimensions["E"].hidden = True

for cell in ["A3", "C3"]:
    dash[cell].font = Font(bold=True, color=green_dark)
    dash[cell].alignment = Alignment(horizontal="left", vertical="center")
for cell in ["B3", "D3"]:
    dash[cell].fill = PatternFill("solid", fgColor=gold_light)
    dash[cell].font = Font(bold=True, color="000000")
    dash[cell].alignment = Alignment(horizontal="center", vertical="center")

# Replace hidden validation lists with user-friendly year/month choices.
for idx, year in enumerate([2569, 2570, 2571, 2572, 2573], 1):
    dash[f"N{idx}"] = year
months = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
for idx, month in enumerate(months, 1):
    dash[f"O{idx}"] = month
dash["P1"] = "ทั้งหมด"
for col in ["N", "O", "P"]:
    dash.column_dimensions[col].hidden = True

dash.data_validations.dataValidation = []
dv_year = DataValidation(type="list", formula1="=$N$1:$N$5", allow_blank=False)
dv_month = DataValidation(type="list", formula1="=$P$1,$O$1:$O$12", allow_blank=False)
dash.add_data_validation(dv_year)
dash.add_data_validation(dv_month)
dv_year.add(dash["B3"])
dv_month.add(dash["D3"])

# Summary helper cells are the single source for the dashboard filter.
summary["Q4"] = "ตัวกรอง Dashboard"
summary["R4"] = "ค่า"
summary["Q5"] = "ปี"
summary["R5"] = "='Dashboard'!B3"
summary["Q6"] = "เดือน"
summary["R6"] = "='Dashboard'!E3"

products = ["ทุเรียน", "มังคุด", "ลองกอง", "ยังไม่ชัว"]
status_criteria = "'ฐานข้อมูล'!$P:$P,\"ปกติ\""
year_criteria = "'ฐานข้อมูล'!$C:$C,$R$5"

summary["Q8"] = "สินค้า"
summary["R8"] = "รายรับ"
summary["S8"] = "รายจ่าย"
summary["T8"] = "สุทธิ"
for idx, product in enumerate(products, 9):
    summary.cell(idx, 17, product)
    product_criteria = f"'ฐานข้อมูล'!$D:$D,Q{idx}"
    start_date = "DATE($R$5-543,$R$6,1)"
    next_month = f"EDATE({start_date},1)"
    revenue_all = f"SUMIFS('ฐานข้อมูล'!$J:$J,{product_criteria},{status_criteria},{year_criteria})"
    expense_all = f"SUMIFS('ฐานข้อมูล'!$K:$K,{product_criteria},{status_criteria},{year_criteria})+SUMIFS('ฐานข้อมูล'!$L:$L,{product_criteria},{status_criteria},{year_criteria})"
    revenue_month = f"SUMIFS('ฐานข้อมูล'!$J:$J,{product_criteria},{status_criteria},{year_criteria},'ฐานข้อมูล'!$B:$B,\">=\"&{start_date},'ฐานข้อมูล'!$B:$B,\"<\"&{next_month})"
    expense_month = f"SUMIFS('ฐานข้อมูล'!$K:$K,{product_criteria},{status_criteria},{year_criteria},'ฐานข้อมูล'!$B:$B,\">=\"&{start_date},'ฐานข้อมูล'!$B:$B,\"<\"&{next_month})+SUMIFS('ฐานข้อมูล'!$L:$L,{product_criteria},{status_criteria},{year_criteria},'ฐานข้อมูล'!$B:$B,\">=\"&{start_date},'ฐานข้อมูล'!$B:$B,\"<\"&{next_month})"
    summary.cell(idx, 18, f'=IF($R$6="ทั้งหมด",{revenue_all},{revenue_month})')
    summary.cell(idx, 19, f'=IF($R$6="ทั้งหมด",{expense_all},{expense_month})')
    summary.cell(idx, 20, f"=R{idx}-S{idx}")

summary["Q15"] = "ชื่อ/รายการ"
summary["R15"] = "รายจ่าย"
for row in range(16, 26):
    name_cell = f"Q{row}"
    start_date = "DATE($R$5-543,$R$6,1)"
    next_month = f"EDATE({start_date},1)"
    base = f"'ฐานข้อมูล'!$G:$G,{name_cell},{status_criteria},{year_criteria}"
    expense_all = f"SUMIFS('ฐานข้อมูล'!$K:$K,{base})+SUMIFS('ฐานข้อมูล'!$L:$L,{base})"
    expense_month = f"SUMIFS('ฐานข้อมูล'!$K:$K,{base},'ฐานข้อมูล'!$B:$B,\">=\"&{start_date},'ฐานข้อมูล'!$B:$B,\"<\"&{next_month})+SUMIFS('ฐานข้อมูล'!$L:$L,{base},'ฐานข้อมูล'!$B:$B,\">=\"&{start_date},'ฐานข้อมูล'!$B:$B,\"<\"&{next_month})"
    summary.cell(row, 18, f'=IF($R$6="ทั้งหมด",{expense_all},{expense_month})')

summary["U8"] = "KPI กรอง"
summary["V8"] = "ค่า"
summary["U9"] = "รายรับ"
summary["V9"] = "=SUM(R9:R12)"
summary["U10"] = "รายจ่าย"
summary["V10"] = "=SUM(S9:S12)"
summary["U11"] = "สุทธิ"
summary["V11"] = "=V9-V10"
summary["U12"] = "กก.รวม"
summary["V12"] = '=IF($R$6="ทั้งหมด",SUMIFS(\'ฐานข้อมูล\'!$I:$I,\'ฐานข้อมูล\'!$C:$C,$R$5,\'ฐานข้อมูล\'!$P:$P,"ปกติ"),SUMIFS(\'ฐานข้อมูล\'!$I:$I,\'ฐานข้อมูล\'!$C:$C,$R$5,\'ฐานข้อมูล\'!$P:$P,"ปกติ",\'ฐานข้อมูล\'!$B:$B,">="&DATE($R$5-543,$R$6,1),\'ฐานข้อมูล\'!$B:$B,"<"&EDATE(DATE($R$5-543,$R$6,1),1)))'
summary["U13"] = "ราคาเฉลี่ย/กก."
summary["V13"] = "=IFERROR(V9/V12,0)"

# Dashboard KPI cards use the filtered helper block.
dash["A5"] = "='สรุป'!V9"
dash["C5"] = "='สรุป'!V10"
dash["E5"] = "='สรุป'!V11"
dash["G5"] = "='สรุป'!V12"
dash["I5"] = "='สรุป'!V13"

# Dashboard tables link to the filtered summary block.
for r in range(10, 14):
    src_row = r - 1
    dash.cell(r, 1, f"='สรุป'!Q{src_row}")
    dash.cell(r, 2, f"='สรุป'!R{src_row}")
    dash.cell(r, 3, f"='สรุป'!S{src_row}")
    dash.cell(r, 4, f"='สรุป'!T{src_row}")

for r in range(10, 20):
    src_row = r + 6
    dash.cell(r, 8, f"='สรุป'!Q{src_row}")
    dash.cell(r, 9, f"='สรุป'!R{src_row}")

# Monthly summary always shows readable month names and links to the summary month table.
dash["A17"] = "สรุปรายเดือน"
for col, header in enumerate(["เดือน", "รายรับ", "รายจ่าย", "สุทธิ"], 1):
    dash.cell(18, col, header)
for r, src_row in [(19, 5), (20, 6)]:
    dash.cell(r, 1, f'=TEXT(\'สรุป\'!M{src_row},"mmm yyyy")')
    dash.cell(r, 2, f"='สรุป'!N{src_row}")
    dash.cell(r, 3, f"='สรุป'!O{src_row}")
    dash.cell(r, 4, f"=B{r}-C{r}")

for row in dash.iter_rows(min_row=5, max_row=21, min_col=1, max_col=9):
    for cell in row:
        if cell.column in [2, 3, 4, 5, 7, 9]:
            cell.number_format = "#,##0.00"

for header_row in [9, 18]:
    for cell in dash[header_row]:
        if cell.value:
            cell.fill = PatternFill("solid", fgColor=green_dark)
            cell.font = Font(color=white, bold=True)
            cell.alignment = Alignment(horizontal="center")

for sheet in wb.worksheets:
    sheet.protection.sheet = False
    sheet.protection.enable = False

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"

wb.save(out)
print(out)
