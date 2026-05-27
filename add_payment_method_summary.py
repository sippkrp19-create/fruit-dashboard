from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo


SRC = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_สรุปมีรายรับ.xlsx")
MAX_ROW = 2000

GREEN = "1F4E3D"
WHITE = "FFFFFF"
GREEN_LIGHT = "E9F4EC"
GRAY_TEXT = "60706A"


def month_filtered_sum(value_col: str, criteria: str = "") -> str:
    base = (
        f"ฐานข้อมูล!${value_col}$5:${value_col}${MAX_ROW},"
        f"ฐานข้อมูล!$O$5:$O${MAX_ROW},\"ปกติ\","
        f"ฐานข้อมูล!$B$5:$B${MAX_ROW},$B$3"
    )
    month = base
    if criteria:
        base += f",{criteria}"
        month += f",{criteria}"
    month += (
        f",ฐานข้อมูล!$A$5:$A${MAX_ROW},\">=\"&DATE($B$3-543,$E$3,1),"
        f"ฐานข้อมูล!$A$5:$A${MAX_ROW},\"<\"&EDATE(DATE($B$3-543,$E$3,1),1)"
    )
    return f'IF($E$3="ทั้งหมด",SUMIFS({base}),SUMIFS({month}))'


def method_sum(value_col: str, criteria_cell: str) -> str:
    criteria_normal = f"ฐานข้อมูล!$N$5:$N${MAX_ROW},{criteria_cell}"
    if criteria_cell == "L6":
        # Count both old and user-facing transfer labels.
        return (
            f"{month_filtered_sum(value_col, 'ฐานข้อมูล!$N$5:$N$' + str(MAX_ROW) + ',\"โอน\"')}"
            f"+{month_filtered_sum(value_col, 'ฐานข้อมูล!$N$5:$N$' + str(MAX_ROW) + ',\"เงินโอน\"')}"
        )
    return month_filtered_sum(value_col, criteria_normal)


def add_payment_summary():
    wb = openpyxl.load_workbook(SRC)
    ws = wb["สรุป"]

    # Clear the target area before rebuilding this section.
    for row in range(3, 11):
        for col in range(12, 16):
            cell = ws.cell(row, col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()

    ws["L3"] = "แยกตามวิธีรับ/จ่ายเงิน"
    ws["L3"].font = Font(bold=True, color=GREEN, size=12)
    ws.merge_cells("L3:O3")

    headers = ["วิธีรับ/จ่ายเงิน", "รายรับ", "รายจ่าย", "สุทธิ"]
    for col, header in enumerate(headers, 12):
        cell = ws.cell(4, col, header)
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.font = Font(bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal="center")

    methods = ["เงินสด", "เงินโอน", "ค้างรับ", "ค้างจ่าย"]
    for row, method in enumerate(methods, 5):
        ws.cell(row, 12, method)
        ws.cell(row, 13, f"={method_sum('I', f'L{row}')}")
        ws.cell(row, 14, f"={method_sum('J', f'L{row}')}+{method_sum('K', f'L{row}')}+{method_sum('L', f'L{row}')}")
        ws.cell(row, 15, f"=M{row}-N{row}")
        for col in range(12, 16):
            cell = ws.cell(row, col)
            cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT if row % 2 == 1 else "FFFFFF")
            cell.border = Border(bottom=Side(style="thin", color="D8E0DC"))
            if col >= 13:
                cell.number_format = "#,##0.00"

    ws.column_dimensions["L"].width = 20
    for col in ["M", "N", "O"]:
        ws.column_dimensions[col].width = 14

    if "tblByPaymentMethod" in ws.tables:
        del ws.tables["tblByPaymentMethod"]
    table = Table(displayName="tblByPaymentMethod", ref="L4:O8")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    # Keep transfer wording easy for future entries.
    if "รายการดรอปดาวน์" in wb.sheetnames:
        lists = wb["รายการดรอปดาวน์"]
        method_values = [lists.cell(row, 6).value for row in range(5, 105)]
        if "เงินโอน" not in method_values:
            for row in range(5, 105):
                if lists.cell(row, 6).value in (None, ""):
                    lists.cell(row, 6, "เงินโอน")
                    break

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(SRC)
    return SRC


if __name__ == "__main__":
    print(add_payment_summary())
