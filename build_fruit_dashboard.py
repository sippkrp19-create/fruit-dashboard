from __future__ import annotations

import datetime as dt
import re
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


SRC_DURIAN = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\ข้อมูลทุเรียน.xlsx")
SRC_MANGOSTEEN = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\รายรับ-จ่ายมังคุด.xlsx")
OUT = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชีและแดชบอร์ด.xlsx")


GREEN_DARK = "1F4E3D"
GREEN = "4F8A5B"
GREEN_LIGHT = "E9F4EC"
GOLD = "D9A441"
GOLD_LIGHT = "FFF3D6"
GRAY_BG = "F5F7F8"
GRAY_TEXT = "60706A"
RED_LIGHT = "FCE4E4"
WHITE = "FFFFFF"
BLACK = "1F2523"
BLUE_LIGHT = "E8F1FA"


HEADERS = [
    "รหัสรายการ",
    "วันที่",
    "รอบ/ฤดูกาล",
    "สินค้า",
    "ประเภทรายการ",
    "สายพันธุ์/เกรด",
    "ชื่อ/รายการ",
    "ราคาต่อกก.",
    "ปริมาณกก.",
    "รายรับ",
    "รายจ่าย",
    "ค่าตัด",
    "เบิก",
    "สุทธิ",
    "วิธีรับ/จ่ายเงิน",
    "สถานะรายการ",
    "วันที่บันทึก",
    "หมายเหตุ",
    "แหล่งที่มา",
]


def clean_name(value):
    if value is None:
        return ""
    return str(value).strip()


def parse_season(value) -> int | None:
    if isinstance(value, dt.datetime):
        year = value.year
        return year + 543 if year < 2500 else year
    if isinstance(value, dt.date):
        year = value.year
        return year + 543 if year < 2500 else year
    text = str(value or "")
    m = re.search(r"(25\d{2}|26\d{2}|\d{2})", text)
    if not m:
        return None
    y = int(m.group(1))
    if y < 100:
        y += 2500
    return y


def parse_date(value, fallback_month=None, fallback_year=2569):
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        y = value.year - 543 if value.year > 2400 else value.year
        return dt.date(y, value.month, value.day)
    if isinstance(value, dt.date):
        y = value.year - 543 if value.year > 2400 else value.year
        return dt.date(y, value.month, value.day)
    text = str(value).strip()
    month_map = {
        "ม.ค": 1,
        "ก.พ": 2,
        "มี.ค": 3,
        "เม.ย": 4,
        "พ.ค": 5,
        "มิ.ย": 6,
        "ก.ค": 7,
        "ส.ค": 8,
        "ก.ย": 9,
        "ต.ค": 10,
        "พ.ย": 11,
        "ธ.ค": 12,
    }
    m = re.search(r"(\d{1,2})\s*([ก-ฮ]\.?\s?[ก-ฮ]?\.?)\s*(\d{2,4})", text)
    if m:
        day = int(m.group(1))
        mon_txt = m.group(2).replace(" ", "")
        month = month_map.get(mon_txt)
        year = int(m.group(3))
        if year < 100:
            year += 2500
        if month:
            return dt.date(year - 543 if year > 2400 else year, month, day)
    m = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", text)
    if m:
        day, month, year = map(int, m.groups())
        if year < 100:
            year += 2500
        return dt.date(year - 543 if year > 2400 else year, month, day)
    if fallback_month:
        m = re.search(r"(\d{1,2})", text)
        if m:
            year = fallback_year
            return dt.date(year - 543 if year > 2400 else year, fallback_month, int(m.group(1)))
    return None


def baht(value):
    return value if isinstance(value, (int, float)) else 0


def extract_deduction(note: str | None):
    if not note:
        return 0, None
    text = str(note)
    m = re.search(r"=\s*([\d,]+(?:\.\d+)?)", text)
    deduction = float(m.group(1).replace(",", "")) if m else 0
    net = None
    m2 = re.search(r"เหลือ\s*([\d,]+(?:\.\d+)?)", text)
    if m2:
        net = float(m2.group(1).replace(",", ""))
    return deduction, net


def load_source_rows():
    rows = []
    wb = load_workbook(SRC_DURIAN, data_only=True)
    ws = wb["ข้อมูลผลผลิตทุเรียน"]
    txn = 1
    for r in range(4, ws.max_row + 1, 3):
        date_value = ws.cell(r, 2).value
        date = parse_date(date_value)
        if not date:
            continue
        season = parse_season(date_value) or 2569
        note = ws.cell(r + 2, 29).value
        deduction, note_net = extract_deduction(note)
        first = True
        for c in range(4, 28):
            grade = clean_name(ws.cell(3, c).value)
            price = ws.cell(r, c).value
            qty = ws.cell(r + 1, c).value
            revenue = ws.cell(r + 2, c).value
            if not any(isinstance(v, (int, float)) and v != 0 for v in [price, qty, revenue]):
                continue
            row_deduction = deduction if first else 0
            row_revenue = baht(revenue)
            rows.append([
                f"TXN-{txn:04d}",
                date,
                season,
                "ทุเรียน",
                "รายรับ",
                grade,
                "",
                price if isinstance(price, (int, float)) else None,
                qty if isinstance(qty, (int, float)) else None,
                None,
                None,
                row_deduction,
                0,
                None,
                "เงินสด",
                "ปกติ",
                dt.date.today(),
                note if first else "",
                f"ต้นฉบับทุเรียน!แถว {r}-{r+2}",
            ])
            txn += 1
            first = False

    wb = load_workbook(SRC_MANGOSTEEN, data_only=True)
    for sheet, month_num in [("รายรับ เดือน เม.ย", 4), ("รายรับ เดือน พ.ค", 5)]:
        ws = wb[sheet]
        for r in range(2, ws.max_row + 1):
            date = parse_date(ws.cell(r, 1).value, fallback_month=month_num)
            price, qty = ws.cell(r, 2).value, ws.cell(r, 3).value
            if not date or not isinstance(qty, (int, float)):
                continue
            rows.append([
                f"TXN-{txn:04d}",
                date,
                parse_season(ws.cell(r, 1).value) or 2569,
                "มังคุด",
                "รายรับ",
                "",
                "",
                price if isinstance(price, (int, float)) else None,
                qty,
                None,
                None,
                0,
                0,
                None,
                "เงินสด",
                "ปกติ",
                dt.date.today(),
                "",
                f"{sheet}!แถว {r}",
            ])
            txn += 1

    for sheet, month_num in [("รายจ่ายเดือน เม.ย", 4), ("รายจ่ายเดือน พ.ค", 5)]:
        ws = wb[sheet]
        current_date = None
        for r in range(1 if month_num == 5 else 2, ws.max_row + 1):
            if ws.cell(r, 1).value:
                current_date = parse_date(ws.cell(r, 1).value, fallback_month=month_num)
            name = clean_name(ws.cell(r, 2).value)
            price, qty, amount = ws.cell(r, 3).value, ws.cell(r, 4).value, ws.cell(r, 5).value
            withdraw = ws.cell(r, 6).value
            if not current_date or not name or not isinstance(qty, (int, float)):
                continue
            rows.append([
                f"TXN-{txn:04d}",
                current_date,
                2569,
                "มังคุด",
                "รายจ่าย",
                "",
                name,
                price if isinstance(price, (int, float)) else None,
                qty,
                None,
                None,
                0,
                withdraw if isinstance(withdraw, (int, float)) else 0,
                None,
                "เงินสด",
                "ปกติ",
                dt.date.today(),
                "",
                f"{sheet}!แถว {r}",
            ])
            txn += 1
    return rows


def set_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def style_title(ws, title, subtitle=""):
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=GREEN_DARK)
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Aptos", size=10, color=GRAY_TEXT)
    ws.freeze_panes = "A4"


def style_table_header(ws, row):
    fill = PatternFill("solid", fgColor=GREEN_DARK)
    for cell in ws[row]:
        if cell.value is not None:
            cell.fill = fill
            cell.font = Font(color=WHITE, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def add_table(ws, ref, name, style="TableStyleMedium4"):
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name=style, showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)


def fill_original_sheet(target_ws, source_path, max_gap=2):
    target_ws.sheet_view.showGridLines = False
    target_ws["A1"] = "ข้อมูลต้นฉบับสำหรับตรวจเทียบ"
    target_ws["A1"].font = Font(size=18, bold=True, color=GREEN_DARK)
    out_row = 3
    wb = load_workbook(source_path, data_only=False)
    for ws in wb.worksheets:
        target_ws.cell(out_row, 1, f"ไฟล์: {source_path.name} | ชีต: {ws.title}")
        target_ws.cell(out_row, 1).font = Font(bold=True, color=WHITE)
        target_ws.cell(out_row, 1).fill = PatternFill("solid", fgColor=GREEN_DARK)
        out_row += 1
        for r in ws.iter_rows(values_only=True):
            for c, v in enumerate(r, 1):
                target_ws.cell(out_row, c, v)
            out_row += 1
        out_row += max_gap
    set_col_widths(target_ws, {"A": 18, "B": 20, "C": 20, "D": 16, "E": 16, "F": 16, "G": 16, "H": 16, "I": 16, "J": 16, "K": 16})


def build():
    rows = load_source_rows()
    expense_totals = {}
    for row in rows:
        name = row[6]
        if row[3] == "มังคุด" and row[4] == "รายจ่าย" and name:
            expense_totals[name] = expense_totals.get(name, 0) + (row[7] or 0) * (row[8] or 0)
    wb = Workbook()
    wb.remove(wb.active)
    dash = wb.create_sheet("Dashboard")
    form = wb.create_sheet("เพิ่มข้อมูล")
    db = wb.create_sheet("ฐานข้อมูล")
    summary = wb.create_sheet("สรุป")
    original = wb.create_sheet("ต้นฉบับ")

    # Database
    style_title(db, "ฐานข้อมูล", "ตารางรวมรายการทั้งหมด ใช้เป็นแหล่งข้อมูลหลักของสรุปและ Dashboard")
    for col_idx, header in enumerate(HEADERS, 1):
        db.cell(4, col_idx, header)
    for row_idx, row in enumerate(rows, 5):
        for col_idx, value in enumerate(row, 1):
            db.cell(row_idx, col_idx, value)
    for excel_row in range(5, len(rows) + 5):
        db.cell(excel_row, 10, f'=IF($E{excel_row}="รายรับ",$H{excel_row}*$I{excel_row},0)')
        db.cell(excel_row, 11, f'=IF($E{excel_row}="รายจ่าย",$H{excel_row}*$I{excel_row},0)')
        db.cell(excel_row, 14, f"=J{excel_row}-K{excel_row}-L{excel_row}")
    style_table_header(db, 4)
    add_table(db, f"A4:S{len(rows)+4}", "tblDatabase")
    db.freeze_panes = "A5"
    set_col_widths(db, {
        "A": 14, "B": 13, "C": 12, "D": 12, "E": 14, "F": 24, "G": 20,
        "H": 12, "I": 12, "J": 14, "K": 14, "L": 12, "M": 12, "N": 14,
        "O": 16, "P": 14, "Q": 14, "R": 36, "S": 22,
    })
    for row in db.iter_rows(min_row=5, max_row=len(rows) + 4):
        row[1].number_format = "d/m/yyyy"
        row[7].number_format = '#,##0.00'
        row[8].number_format = '#,##0.00'
        for idx in [9, 10, 11, 12, 13]:
            row[idx].number_format = '#,##0.00'
        row[16].number_format = "d/m/yyyy"
    db.auto_filter.ref = f"A4:S{len(rows)+4}"

    # Hidden lists on form
    products = ["ทุเรียน", "มังคุด", "ลองกอง", "ยังไม่ชัว"]
    types = ["รายรับ", "รายจ่าย", "ผลผลิต"]
    pay_methods = ["เงินสด", "โอน", "ค้างรับ", "ค้างจ่าย"]
    statuses = ["ปกติ", "ยกเลิก", "แก้ไขแล้ว"]
    grades = sorted({r[5] for r in rows if r[5]})
    names = [name for name, _ in sorted(expense_totals.items(), key=lambda item: item[1], reverse=True)]
    seasons = sorted({r[2] for r in rows if r[2]})
    lists = {
        "AA": products,
        "AB": types,
        "AC": grades,
        "AD": names,
        "AE": pay_methods,
        "AF": statuses,
        "AG": seasons,
    }

    # Form
    style_title(form, "เพิ่มข้อมูล", "กรอกข้อมูลรายการใหม่ ตรวจสถานะ แล้วคัดลอกแถวตัวอย่างไปเพิ่มท้ายตารางฐานข้อมูล")
    set_col_widths(form, {"A": 4, "B": 22, "C": 24, "D": 4, "E": 22, "F": 24, "H": 18, "I": 18, "J": 18, "K": 18, "L": 18, "M": 18})
    for col, values in lists.items():
        form[f"{col}1"] = col
        for i, value in enumerate(values, 2):
            form[f"{col}{i}"] = value
        form.column_dimensions[col].hidden = True

    input_fill = PatternFill("solid", fgColor=GOLD_LIGHT)
    calc_fill = PatternFill("solid", fgColor=GRAY_BG)
    label_font = Font(bold=True, color=GREEN_DARK)
    fields = [
        ("B4", "วันที่", "C4"),
        ("B5", "รอบ/ฤดูกาล", "C5"),
        ("B6", "สินค้า", "C6"),
        ("B7", "ประเภทรายการ", "C7"),
        ("B8", "สายพันธุ์/เกรด", "C8"),
        ("B9", "ชื่อ/รายการ", "C9"),
        ("B10", "ราคาต่อกก.", "C10"),
        ("B11", "ปริมาณกก.", "C11"),
        ("E4", "ค่าตัด", "F4"),
        ("E5", "เบิก", "F5"),
        ("E6", "วิธีรับ/จ่ายเงิน", "F6"),
        ("E7", "สถานะรายการ", "F7"),
        ("E8", "หมายเหตุ", "F8"),
        ("E10", "รายรับ", "F10"),
        ("E11", "รายจ่าย", "F11"),
        ("E12", "สุทธิ", "F12"),
        ("E14", "สถานะตรวจสอบ", "F14"),
    ]
    for label_cell, label, value_cell in fields:
        form[label_cell] = label
        form[label_cell].font = label_font
        form[value_cell].fill = input_fill if value_cell not in ["F10", "F11", "F12", "F14"] else calc_fill
        form[value_cell].border = Border(bottom=Side(style="thin", color="D8E0DC"))
    form["C4"] = dt.date.today()
    form["C4"].number_format = "d/m/yyyy"
    form["C5"] = 2569
    form["F4"] = 0
    form["F5"] = 0
    form["F7"] = "ปกติ"
    form["F10"] = '=IF($C$7="รายรับ",$C$10*$C$11,0)'
    form["F11"] = '=IF($C$7="รายจ่าย",$C$10*$C$11,0)'
    form["F12"] = '=F10-F11-F4'
    form["F14"] = '=IF(OR(C4="",C5="",C6="",C7="",C11=""),"กรอกข้อมูลไม่ครบ",IF(OR(C10="",C10<0,C11<=0),"ตรวจสอบจำนวน/ราคา","พร้อมบันทึก"))'
    form["B16"] = "แถวตัวอย่างสำหรับเพิ่มท้ายฐานข้อมูล"
    form["B16"].font = Font(bold=True, color=GREEN_DARK)
    preview_headers = HEADERS
    for idx, header in enumerate(preview_headers, 2):
        cell = form.cell(17, idx, header)
        cell.fill = PatternFill("solid", fgColor=GREEN_DARK)
        cell.font = Font(color=WHITE, bold=True)
    preview_formulas = [
        '=TEXT(COUNTA(\'ฐานข้อมูล\'!A:A)-3+1,"TXN-0000")',
        '=C4','=C5','=C6','=C7','=C8','=C9','=C10','=C11','=F10','=F11','=F4','=F5','=F12','=F6','=F7','=TODAY()','=F8','"เพิ่มข้อมูล"',
    ]
    for idx, formula in enumerate(preview_formulas, 2):
        form.cell(18, idx, formula)
    form["B20"] = "วิธีใช้: กรอกช่องสีเหลือง ตรวจให้สถานะเป็น “พร้อมบันทึก” แล้วคัดลอกแถว 18 ไปเพิ่มท้ายตารางในชีตฐานข้อมูล"
    form["B20"].font = Font(color=GRAY_TEXT, italic=True)
    form.merge_cells("B20:T20")
    for col, values in lists.items():
        max_row = len(values) + 1
        dv_range = f"${col}$2:${col}${max_row}"
        target = {"AA": "C6", "AB": "C7", "AC": "C8", "AD": "C9", "AE": "F6", "AF": "F7", "AG": "C5"}[col]
        dv = DataValidation(type="list", formula1=f"={dv_range}", allow_blank=True)
        form.add_data_validation(dv)
        dv.add(form[target])
    for c in ["C10", "C11", "F4", "F5", "F10", "F11", "F12"]:
        form[c].number_format = '#,##0.00'
    form["F14"].font = Font(bold=True, color=GREEN_DARK)
    form.conditional_formatting.add("F14", CellIsRule(operator="notEqual", formula=['"พร้อมบันทึก"'], fill=PatternFill("solid", fgColor=RED_LIGHT)))
    form.protection.sheet = True
    for cell in ["C4", "C5", "C6", "C7", "C8", "C9", "C10", "C11", "F4", "F5", "F6", "F7", "F8"]:
        form[cell].protection = Protection(locked=False)

    # Summary
    style_title(summary, "สรุป", "ตารางคำนวณกลางสำหรับ Dashboard นับเฉพาะรายการสถานะปกติ")
    set_col_widths(summary, {"A": 22, "B": 18, "C": 18, "D": 18, "E": 18, "G": 18, "H": 16, "I": 16, "J": 16, "K": 16, "M": 20, "N": 16})
    summary["A4"] = "KPI"
    summary["B4"] = "ค่า"
    db_status = "'ฐานข้อมูล'!$P:$P"
    kpis = [
        ("รายรับรวม", f'=SUMIFS(\'ฐานข้อมูล\'!$J:$J,{db_status},"ปกติ")'),
        ("รายจ่ายรวม", f'=SUMIFS(\'ฐานข้อมูล\'!$K:$K,{db_status},"ปกติ")'),
        ("กำไรสุทธิ", '=B5-B6'),
        ("ปริมาณกก.รวม", f'=SUMIFS(\'ฐานข้อมูล\'!$I:$I,{db_status},"ปกติ")'),
        ("ราคาเฉลี่ย/กก.", '=IFERROR(B5/B8,0)'),
        ("ค่าตัดรวม", f'=SUMIFS(\'ฐานข้อมูล\'!$L:$L,{db_status},"ปกติ")'),
        ("เบิกรวม", f'=SUMIFS(\'ฐานข้อมูล\'!$M:$M,{db_status},"ปกติ")'),
    ]
    for i, (label, formula) in enumerate(kpis, 5):
        summary.cell(i, 1, label)
        summary.cell(i, 2, formula)
        summary.cell(i, 2).number_format = '#,##0.00'
    style_table_header(summary, 4)
    add_table(summary, "A4:B11", "tblKPI", "TableStyleMedium4")
    summary["D4"] = "สินค้า"
    summary["E4"] = "รายรับ"
    summary["F4"] = "รายจ่าย"
    summary["G4"] = "สุทธิ"
    for idx, product in enumerate(products, 5):
        summary.cell(idx, 4, product)
        summary.cell(idx, 5, f'=SUMIFS(\'ฐานข้อมูล\'!$J:$J,\'ฐานข้อมูล\'!$D:$D,D{idx},{db_status},"ปกติ")')
        summary.cell(idx, 6, f'=SUMIFS(\'ฐานข้อมูล\'!$K:$K,\'ฐานข้อมูล\'!$D:$D,D{idx},{db_status},"ปกติ")+SUMIFS(\'ฐานข้อมูล\'!$L:$L,\'ฐานข้อมูล\'!$D:$D,D{idx},{db_status},"ปกติ")')
        summary.cell(idx, 7, f'=E{idx}-F{idx}')
    style_table_header(summary, 4)
    add_table(summary, f"D4:G{4+len(products)}", "tblByProduct", "TableStyleMedium4")
    summary["I4"] = "ชื่อ/รายการ"
    summary["J4"] = "รายจ่าย"
    for idx, name in enumerate(names[:15], 5):
        summary.cell(idx, 9, name)
        summary.cell(idx, 10, f'=SUMIFS(\'ฐานข้อมูล\'!$K:$K,\'ฐานข้อมูล\'!$G:$G,I{idx},{db_status},"ปกติ")')
    style_table_header(summary, 4)
    add_table(summary, f"I4:J{4+len(names[:15])}", "tblExpenseByName", "TableStyleMedium4")
    summary["M4"] = "เดือน"
    summary["N4"] = "รายรับ"
    summary["O4"] = "รายจ่าย"
    for idx, month in enumerate([4, 5], 5):
        summary.cell(idx, 13, dt.date(2026, month, 1))
        summary.cell(idx, 13).number_format = "mmm yyyy"
        summary.cell(idx, 14, f'=SUMIFS(\'ฐานข้อมูล\'!$J:$J,\'ฐานข้อมูล\'!$B:$B,">="&M{idx},\'ฐานข้อมูล\'!$B:$B,"<"&EDATE(M{idx},1),{db_status},"ปกติ")')
        summary.cell(idx, 15, f'=SUMIFS(\'ฐานข้อมูล\'!$K:$K,\'ฐานข้อมูล\'!$B:$B,">="&M{idx},\'ฐานข้อมูล\'!$B:$B,"<"&EDATE(M{idx},1),{db_status},"ปกติ")+SUMIFS(\'ฐานข้อมูล\'!$L:$L,\'ฐานข้อมูล\'!$B:$B,">="&M{idx},\'ฐานข้อมูล\'!$B:$B,"<"&EDATE(M{idx},1),{db_status},"ปกติ")')
    add_table(summary, "M4:O6", "tblByMonth", "TableStyleMedium4")
    # Dashboard-filtered calculation block.
    summary["Q4"] = "ตัวกรอง Dashboard"
    summary["R4"] = "ค่า"
    summary["Q5"] = "ปี"
    summary["R5"] = "='Dashboard'!B3"
    summary["Q6"] = "เดือน"
    summary["R6"] = "='Dashboard'!D3"
    summary["Q8"] = "สินค้า"
    summary["R8"] = "รายรับ"
    summary["S8"] = "รายจ่าย"
    summary["T8"] = "สุทธิ"
    for idx, product in enumerate(products, 9):
        summary.cell(idx, 17, product)
        base = f'\'ฐานข้อมูล\'!$D:$D,Q{idx},{db_status},"ปกติ"'
        year_crit = '\'ฐานข้อมูล\'!$C:$C,$R$5'
        month_rev = f'IF($R$6="ทั้งหมด",SUMIFS(\'ฐานข้อมูล\'!$J:$J,{base},{year_crit}),SUMIFS(\'ฐานข้อมูล\'!$J:$J,{base},{year_crit},\'ฐานข้อมูล\'!$B:$B,">="&DATE($R$5-543,$R$6,1),\'ฐานข้อมูล\'!$B:$B,"<"&EDATE(DATE($R$5-543,$R$6,1),1)))'
        month_exp = f'IF($R$6="ทั้งหมด",SUMIFS(\'ฐานข้อมูล\'!$K:$K,{base},{year_crit})+SUMIFS(\'ฐานข้อมูล\'!$L:$L,{base},{year_crit}),SUMIFS(\'ฐานข้อมูล\'!$K:$K,{base},{year_crit},\'ฐานข้อมูล\'!$B:$B,">="&DATE($R$5-543,$R$6,1),\'ฐานข้อมูล\'!$B:$B,"<"&EDATE(DATE($R$5-543,$R$6,1),1))+SUMIFS(\'ฐานข้อมูล\'!$L:$L,{base},{year_crit},\'ฐานข้อมูล\'!$B:$B,">="&DATE($R$5-543,$R$6,1),\'ฐานข้อมูล\'!$B:$B,"<"&EDATE(DATE($R$5-543,$R$6,1),1)))'
        summary.cell(idx, 18, f"={month_rev}")
        summary.cell(idx, 19, f"={month_exp}")
        summary.cell(idx, 20, f"=R{idx}-S{idx}")
    summary["Q15"] = "ชื่อ/รายการ"
    summary["R15"] = "รายจ่าย"
    for out_idx, name in enumerate(names[:10], 16):
        summary.cell(out_idx, 17, name)
        base = f'\'ฐานข้อมูล\'!$G:$G,Q{out_idx},{db_status},"ปกติ",\'ฐานข้อมูล\'!$C:$C,$R$5'
        formula = f'=IF($R$6="ทั้งหมด",SUMIFS(\'ฐานข้อมูล\'!$K:$K,{base})+SUMIFS(\'ฐานข้อมูล\'!$L:$L,{base}),SUMIFS(\'ฐานข้อมูล\'!$K:$K,{base},\'ฐานข้อมูล\'!$B:$B,">="&DATE($R$5-543,$R$6,1),\'ฐานข้อมูล\'!$B:$B,"<"&EDATE(DATE($R$5-543,$R$6,1),1))+SUMIFS(\'ฐานข้อมูล\'!$L:$L,{base},\'ฐานข้อมูล\'!$B:$B,">="&DATE($R$5-543,$R$6,1),\'ฐานข้อมูล\'!$B:$B,"<"&EDATE(DATE($R$5-543,$R$6,1),1)))'
        summary.cell(out_idx, 18, formula)
    summary["U8"] = "KPI กรอง"
    summary["V8"] = "ค่า"
    summary["U9"] = "รายรับ"
    summary["V9"] = "=SUM(R9:R12)"
    summary["U10"] = "รายจ่าย"
    summary["V10"] = "=SUM(S9:S12)"
    summary["U11"] = "สุทธิ"
    summary["V11"] = "=V9-V10"
    summary["U12"] = "กก.รวม"
    summary["V12"] = '=IF($R$6="ทั้งหมด",SUMIFS(\'ฐานข้อมูล\'!$I:$I,\'ฐานข้อมูล\'!$C:$C,$R$5,\'ฐานข้อมูล\'!$P:$P,"ปกติ"),SUMIFS(\'ฐานข้อมูล\'!$I:$I,\'ฐานข้อมูล\'!$C:$C,$R$5,\'ฐานข้อมูล\'!$P:$P,"ปกติ",\'ฐานข้อมูล\'!$B:$B,">="&DATE($R$5-543,$R$6,1),\'ฐานข้อมูล\'!$B:$B,"<"&EDATE(DATE($R$5-543,$R$6,1),1)))'
    summary["U13"] = "ราคาเฉลี่ย/กก."
    summary["V13"] = "=IFERROR(V9/V12,0)"
    style_table_header(summary, 8)
    style_table_header(summary, 15)
    for row in summary.iter_rows(min_row=9, max_row=26, min_col=17, max_col=22):
        for cell in row:
            if cell.column in [18, 19, 20, 22]:
                cell.number_format = '#,##0.00'
    for row in summary.iter_rows(min_row=5, max_row=30):
        for cell in row:
            if cell.column in [2, 5, 6, 7, 10, 14, 15]:
                cell.number_format = '#,##0.00'

    # Dashboard
    dash.sheet_view.showGridLines = False
    dash["A1"] = "สวนผลไม้: บัญชีและแดชบอร์ด"
    dash["A1"].font = Font(name="Aptos Display", size=24, bold=True, color=GREEN_DARK)
    dash["A2"] = "ภาพรวมรายรับ รายจ่าย กำไร และผลผลิต จากฐานข้อมูลเดียว"
    dash["A2"].font = Font(size=10, color=GRAY_TEXT)
    set_col_widths(dash, {c: 14 for c in "ABCDEFGHIJKLM"})
    dash["A3"] = "ปี"
    dash["B3"] = 2569
    dash["C3"] = "เดือน"
    dash["D3"] = "ทั้งหมด"
    for cell in ["A3", "C3"]:
        dash[cell].font = Font(bold=True, color=GREEN_DARK)
    for cell in ["B3", "D3"]:
        dash[cell].fill = PatternFill("solid", fgColor=GOLD_LIGHT)
        dash[cell].border = Border(bottom=Side(style="thin", color="D8E0DC"))
    for col, values in {"N": seasons, "O": ["ทั้งหมด", 4, 5]}.items():
        for i, value in enumerate(values, 1):
            dash[f"{col}{i}"] = value
        dash.column_dimensions[col].hidden = True
    dv_year = DataValidation(type="list", formula1=f"=$N$1:$N${len(seasons)}", allow_blank=False)
    dv_month = DataValidation(type="list", formula1="=$O$1:$O$3", allow_blank=False)
    dash.add_data_validation(dv_year)
    dash.add_data_validation(dv_month)
    dv_year.add(dash["B3"])
    dv_month.add(dash["D3"])
    dash.row_dimensions[4].height = 34
    cards = [
        ("A4", "รายรับ", "='สรุป'!V9", GREEN_DARK),
        ("C4", "รายจ่าย", "='สรุป'!V10", "8A3B35"),
        ("E4", "สุทธิ", "='สรุป'!V11", GREEN),
        ("G4", "กก.รวม", "='สรุป'!V12", GOLD),
        ("I4", "ราคาเฉลี่ย/กก.", "='สรุป'!V13", "34699A"),
    ]
    for cell, label, formula, color in cards:
        col = openpyxl.utils.column_index_from_string(re.sub(r"\d", "", cell))
        dash[cell] = label
        dash[cell].fill = PatternFill("solid", fgColor=color)
        dash[cell].font = Font(color=WHITE, bold=True, size=11)
        dash[cell].alignment = Alignment(horizontal="center")
        val_cell = dash.cell(5, col)
        val_cell.value = formula
        val_cell.fill = PatternFill("solid", fgColor=WHITE)
        val_cell.font = Font(color=BLACK, bold=True, size=16)
        val_cell.alignment = Alignment(horizontal="center")
        val_cell.number_format = '#,##0.00'
        dash.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
        dash.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col+1)
    dash["A8"] = "รายรับเทียบรายจ่ายตามเดือน"
    dash["A8"].font = Font(bold=True, color=GREEN_DARK, size=13)
    dash["A8"] = "สรุปตามสินค้า (ตามตัวกรอง)"
    dash["H8"] = "Top รายจ่ายตามชื่อ/รายการ"
    dash["A17"] = "สรุปรายเดือน"
    for cell in ["A8", "H8", "A17"]:
        dash[cell].font = Font(bold=True, color=GREEN_DARK, size=14)
    product_headers = ["สินค้า", "รายรับ", "รายจ่าย", "สุทธิ"]
    for c, header in enumerate(product_headers, 1):
        dash.cell(9, c, header)
    for r in range(10, 14):
        dash.cell(r, 1, f"='สรุป'!Q{r-1}")
        dash.cell(r, 2, f"='สรุป'!R{r-1}")
        dash.cell(r, 3, f"='สรุป'!S{r-1}")
        dash.cell(r, 4, f"='สรุป'!T{r-1}")
    expense_headers = ["ชื่อ/รายการ", "รายจ่าย"]
    for c, header in enumerate(expense_headers, 8):
        dash.cell(9, c, header)
    for r in range(10, 20):
        dash.cell(r, 8, f"='สรุป'!Q{r+6}")
        dash.cell(r, 9, f"='สรุป'!R{r+6}")
    month_headers = ["เดือน", "รายรับ", "รายจ่าย", "สุทธิ"]
    for c, header in enumerate(month_headers, 1):
        dash.cell(18, c, header)
    for r in range(19, 21):
        src = r - 14
        dash.cell(r, 1, f"='สรุป'!M{src}")
        dash.cell(r, 2, f"='สรุป'!N{src}")
        dash.cell(r, 3, f"='สรุป'!O{src}")
        dash.cell(r, 4, f"=B{r}-C{r}")
    for header_row in [9, 18]:
        for cell in dash[header_row]:
            if cell.value is not None:
                cell.fill = PatternFill("solid", fgColor=GREEN_DARK)
                cell.font = Font(color=WHITE, bold=True)
                cell.alignment = Alignment(horizontal="center")
    for row in dash.iter_rows(min_row=10, max_row=21, min_col=1, max_col=9):
        for cell in row:
            if cell.column in [2, 3, 4, 9]:
                cell.number_format = '#,##0.00'
            if cell.row % 2 == 0 and cell.value is not None:
                cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)

    fill_original_sheet(original, SRC_DURIAN)
    fill_original_sheet(original, SRC_MANGOSTEEN)
    original.sheet_properties.tabColor = "808080"

    for ws in [dash, summary, db, form]:
        ws.sheet_properties.tabColor = GREEN_DARK
        for row in ws.iter_rows():
            for cell in row:
                cell.font = copy(cell.font)
                cell.font = Font(name="Aptos", size=cell.font.sz or 11, bold=cell.font.b, italic=cell.font.i, color=cell.font.color)
                cell.alignment = copy(cell.alignment)
                if cell.alignment.horizontal is None:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

    summary.protection.sheet = True
    dash.protection.sheet = True
    original.protection.sheet = True
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"

    wb.save(OUT)
    return OUT, len(rows)


if __name__ == "__main__":
    out, count = build()
    print(f"saved={out}")
    print(f"rows={count}")
