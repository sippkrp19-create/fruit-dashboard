from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
import tempfile
import urllib.request
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


GOOGLE_SHEET_URL = "https://docs.google.com/spreadsheets/d/1wqsqJkXGchsUCg2mgY06qjKQLoZR4wBLL4f3fjJc-Uc/edit?usp=sharing"
DEFAULT_TARGET = Path(
    r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_สรุปมีรายรับ.xlsx"
)
DEFAULT_OUTPUT = Path(
    r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_ดึงข้อมูลอัตโนมัติ.xlsx"
)

DB_SHEET = "ฐานข้อมูล"
SETUP_SHEET = "ดึงข้อมูลอัตโนมัติ"
HEADER_ROW = 4
DATA_START_ROW = 5
MAX_DATA_ROW = 2000

HEADERS = [
    "วันที่",
    "รอบ/ฤดูกาล",
    "สินค้า",
    "ประเภทรายการ",
    "สายพันธุ์/เกรด",
    "ชื่อ/รายการ",
    "ราคาต่อกก.",
    "ปริมาณกก.",
    "รายรับ",
    "รายจ่าย",
    "ค่าตัด",
    "เบิก",
    "สุทธิ",
    "วิธีรับ/จ่ายเงิน",
    "สถานะรายการ",
    "วันที่บันทึก",
    "หมายเหตุ",
    "แหล่งที่มา",
]


def spreadsheet_id_from_url(url: str) -> str:
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if match:
        return match.group(1)
    if re.fullmatch(r"[a-zA-Z0-9_-]+", url):
        return url
    raise ValueError("ไม่พบ spreadsheet id จากลิงก์ Google Sheet")


def download_google_sheet(url: str) -> Path:
    spreadsheet_id = spreadsheet_id_from_url(url)
    export_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx"
    target = Path(tempfile.gettempdir()) / f"google_sheet_{spreadsheet_id}.xlsx"
    request = urllib.request.Request(export_url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(request, timeout=60) as response:
        data = response.read()
    if not data.startswith(b"PK"):
        raise RuntimeError(
            "ดาวน์โหลด Google Sheet ไม่ได้เป็นไฟล์ Excel อาจต้องตั้งค่าแชร์เป็น Anyone with the link หรือเชื่อมสิทธิ์ Google ใหม่"
        )
    target.write_bytes(data)
    return target


def normalize(value) -> str:
    return str(value or "").strip()


def find_database_sheet(wb: openpyxl.Workbook):
    required = {"วันที่", "สินค้า", "ประเภทรายการ"}
    best = None
    best_score = -1
    for ws in wb.worksheets:
        for row in range(1, min(ws.max_row, 15) + 1):
            values = [normalize(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 40) + 1)]
            score = sum(1 for header in HEADERS if header in values)
            if required.issubset(values) and score > best_score:
                best = (ws, row, values)
                best_score = score
    if not best:
        raise RuntimeError("หาแท็บต้นทางที่มีหัวตารางฐานข้อมูลไม่พบ")
    return best


def row_has_data(values: list) -> bool:
    return any(value not in (None, "") for value in values)


def read_source_rows(source_path: Path) -> list[list]:
    wb = openpyxl.load_workbook(source_path, data_only=False)
    ws, header_row, source_headers = find_database_sheet(wb)
    index_by_header = {header: idx + 1 for idx, header in enumerate(source_headers) if header}
    rows: list[list] = []
    for row in range(header_row + 1, ws.max_row + 1):
        values = []
        for header in HEADERS:
            col = index_by_header.get(header)
            values.append(ws.cell(row, col).value if col else None)
        if row_has_data(values):
            rows.append(values)
    return rows


def clear_database(ws):
    for row in range(DATA_START_ROW, max(ws.max_row, DATA_START_ROW) + 1):
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row, col).value = None


def apply_database_formulas(ws, row: int):
    if ws.cell(row, 9).value in (None, ""):
        ws.cell(row, 9, f'=IF($D{row}="รายรับ",$G{row}*$H{row},0)')
    if ws.cell(row, 10).value in (None, ""):
        ws.cell(row, 10, f'=IF($D{row}="รายจ่าย",$G{row}*$H{row},0)')
    if ws.cell(row, 13).value in (None, ""):
        ws.cell(row, 13, f"=I{row}-J{row}-K{row}")
    if ws.cell(row, 15).value in (None, ""):
        ws.cell(row, 15, "ปกติ")
    if ws.cell(row, 16).value in (None, ""):
        ws.cell(row, 16, dt.date.today())
    if ws.cell(row, 18).value in (None, ""):
        ws.cell(row, 18, "Google Sheet")


def refresh_database_table(ws, last_row: int):
    ws.tables.clear()
    ref = f"A{HEADER_ROW}:R{last_row}"
    table = Table(displayName="tblDatabase", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.auto_filter.ref = ref
    ws.freeze_panes = "A5"


def add_or_refresh_setup_sheet(wb: openpyxl.Workbook, url: str, row_count: int):
    if SETUP_SHEET in wb.sheetnames:
        ws = wb[SETUP_SHEET]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(SETUP_SHEET, 0)

    green = "1F4E3D"
    light = "E9F4EC"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "ดึงข้อมูลอัตโนมัติ"
    ws["A1"].font = Font(size=18, bold=True, color=green)
    ws["A3"] = "Google Sheet"
    ws["B3"] = url
    ws["A4"] = "อัปเดตล่าสุด"
    ws["B4"] = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ws["A5"] = "จำนวนแถวที่นำเข้า"
    ws["B5"] = row_count
    ws["A7"] = "วิธีใช้"
    ws["B7"] = "รันไฟล์ sync_google_sheet_to_dashboard.ps1 เพื่อดึงข้อมูลล่าสุดเข้าแท็บฐานข้อมูล"
    ws["B8"] = "ถ้าดาวน์โหลดไม่ได้ ให้ตั้งค่าแชร์ Google Sheet เป็น Anyone with the link หรือเชื่อมสิทธิ์ Google Drive ใหม่"
    for row in range(3, 9):
        ws.cell(row, 1).font = Font(bold=True, color=green)
        ws.cell(row, 2).fill = PatternFill("solid", fgColor=light)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 110
    ws["B3"].hyperlink = url
    ws["B3"].style = "Hyperlink"


def sync(source_url: str, target_path: Path, output_path: Path | None = None) -> Path:
    downloaded = download_google_sheet(source_url)
    rows = read_source_rows(downloaded)
    if not rows:
        raise RuntimeError("Google Sheet ไม่มีข้อมูลสำหรับนำเข้า")

    out = output_path or target_path
    wb = openpyxl.load_workbook(target_path)
    if DB_SHEET not in wb.sheetnames:
        raise RuntimeError(f"ไม่พบแท็บ {DB_SHEET!r} ในไฟล์ปลายทาง")
    ws = wb[DB_SHEET]

    for col, header in enumerate(HEADERS, 1):
        ws.cell(HEADER_ROW, col, header)

    clear_database(ws)
    for row_offset, values in enumerate(rows, DATA_START_ROW):
        for col, value in enumerate(values, 1):
            ws.cell(row_offset, col, value)
        apply_database_formulas(ws, row_offset)
        ws.cell(row_offset, 1).number_format = "d/m/yyyy"
        ws.cell(row_offset, 16).number_format = "d/m/yyyy"
        for col in range(7, 14):
            ws.cell(row_offset, col).number_format = "#,##0.00"

    last_row = max(DATA_START_ROW, DATA_START_ROW + len(rows) - 1)
    refresh_database_table(ws, last_row)
    add_or_refresh_setup_sheet(wb, source_url, len(rows))
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(out)
    return out


def main() -> int:
    parser = argparse.ArgumentParser(description="Sync Google Sheet rows into the fruit dashboard workbook.")
    parser.add_argument("--source", default=GOOGLE_SHEET_URL, help="Google Sheet URL or spreadsheet id")
    parser.add_argument("--target", default=str(DEFAULT_TARGET), help="Dashboard workbook to update")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT), help="Optional output workbook path")
    args = parser.parse_args()

    try:
        out = sync(args.source, Path(args.target), Path(args.output) if args.output else None)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1
    print(f"saved={out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
