from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill


src = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_ปรับประสิทธิภาพ.xlsx")
out = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_ปรับฟอร์ม.xlsx")

wb = openpyxl.load_workbook(src)
form = wb["เพิ่มข้อมูล"]

# Ensure the form date defaults to today's date every time the file recalculates.
form["C4"] = "=TODAY()"
form["C4"].number_format = "d/m/yyyy"

# Preview row formulas. A18:S18 mirrors the database columns.
# ID is generated from the current number of database rows, avoiding VALUE errors.
form["A18"] = '=TEXT(COUNTA(\'ฐานข้อมูล\'!$A:$A)-3+1,"TXN-0000")'
form["B18"] = "=C4"
form["C18"] = "=C5"
form["D18"] = "=C6"
form["E18"] = "=C7"
form["F18"] = '=IF(C8<>"",C8,"")'
form["G18"] = '=IF(C9<>"",C9,"")'
form["H18"] = "=C10"
form["I18"] = "=C11"
form["J18"] = "=F10"
form["K18"] = "=F11"
form["L18"] = "=F4"
form["M18"] = "=F5"
form["N18"] = "=F12"
form["O18"] = "=F6"
form["P18"] = "=F7"
form["Q18"] = "=TODAY()"
form["R18"] = "=F8"
form["S18"] = '"เพิ่มข้อมูล"'

for cell in ["B18", "Q18"]:
    form[cell].number_format = "d/m/yyyy"
for cell in ["H18", "I18", "J18", "K18", "L18", "M18", "N18"]:
    form[cell].number_format = "#,##0.00"

# Make calculated preview cells distinct but still editable if needed.
fill = PatternFill("solid", fgColor="E9F4EC")
for row in form.iter_rows(min_row=18, max_row=18, min_col=1, max_col=19):
    for cell in row:
        cell.fill = fill

for sheet in wb.worksheets:
    sheet.protection.sheet = False
    sheet.protection.enable = False

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"

wb.save(out)
print(out)
