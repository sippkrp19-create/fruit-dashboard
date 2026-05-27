from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


src = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_ปรับฟอร์ม.xlsx")
out = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_ฟอร์มตรง.xlsx")

wb = openpyxl.load_workbook(src)
ws = wb["เพิ่มข้อมูล"]

green = "1F4E3D"
light_green = "E9F4EC"
yellow = "FFF3D6"
gray = "F5F7F8"
white = "FFFFFF"

# Rebuild only the visible form area so the preview row cannot drift.
for merged in list(ws.merged_cells.ranges):
    if merged.min_row <= 24 and merged.min_col <= 19:
        ws.unmerge_cells(str(merged))
for row in ws.iter_rows(min_row=1, max_row=24, min_col=1, max_col=19):
    for cell in row:
        cell.value = None
        cell.fill = PatternFill(fill_type=None)
        cell.border = Border()
        cell.font = Font(name="Aptos", size=11)
        cell.alignment = Alignment(horizontal="left", vertical="center")

ws.sheet_view.showGridLines = False
ws.freeze_panes = "A17"

ws["A1"] = "เพิ่มข้อมูล"
ws["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=green)
ws["A2"] = "กรอกช่องสีเหลือง ตรวจสถานะ แล้วคัดลอกแถว 18 ไปเพิ่มท้ายตารางฐานข้อมูล"
ws["A2"].font = Font(italic=True, color="60706A")

fields = [
    ("B4", "วันที่", "C4", "=TODAY()", "d/m/yyyy"),
    ("B5", "รอบ/ฤดูกาล", "C5", 2569, "0"),
    ("B6", "สินค้า", "C6", None, None),
    ("B7", "ประเภทรายการ", "C7", None, None),
    ("B8", "สายพันธุ์/เกรด", "C8", None, None),
    ("B9", "ชื่อ/รายการ", "C9", None, None),
    ("B10", "ราคาต่อกก.", "C10", None, "#,##0.00"),
    ("B11", "ปริมาณกก.", "C11", None, "#,##0.00"),
    ("E4", "ค่าตัด", "F4", 0, "#,##0.00"),
    ("E5", "เบิก", "F5", 0, "#,##0.00"),
    ("E6", "วิธีรับ/จ่ายเงิน", "F6", None, None),
    ("E7", "สถานะรายการ", "F7", "ปกติ", None),
    ("E8", "หมายเหตุ", "F8", None, None),
    ("E10", "รายรับ", "F10", '=IF($C$7="รายรับ",$C$10*$C$11,0)', "#,##0.00"),
    ("E11", "รายจ่าย", "F11", '=IF($C$7="รายจ่าย",$C$10*$C$11,0)', "#,##0.00"),
    ("E12", "สุทธิ", "F12", "=F10-F11-F4", "#,##0.00"),
    ("E14", "สถานะตรวจสอบ", "F14", '=IF(OR(C4="",C5="",C6="",C7="",C11=""),"กรอกข้อมูลไม่ครบ",IF(OR(C10="",C10<0,C11<=0),"ตรวจสอบจำนวน/ราคา","พร้อมบันทึก"))', None),
]

input_fill = PatternFill("solid", fgColor=yellow)
calc_fill = PatternFill("solid", fgColor=gray)
thin = Side(style="thin", color="D8E0DC")

for label_cell, label, value_cell, value, number_format in fields:
    ws[label_cell] = label
    ws[label_cell].font = Font(bold=True, color=green)
    ws[value_cell] = value
    ws[value_cell].fill = calc_fill if value_cell in ["F10", "F11", "F12", "F14"] else input_fill
    ws[value_cell].border = Border(bottom=thin)
    if number_format:
        ws[value_cell].number_format = number_format

ws["A16"] = "แถวตัวอย่างสำหรับเพิ่มท้ายฐานข้อมูล"
ws["A16"].font = Font(bold=True, color=green)

headers = [
    "รหัสรายการ", "วันที่", "รอบ/ฤดูกาล", "สินค้า", "ประเภทรายการ", "สายพันธุ์/เกรด",
    "ชื่อ/รายการ", "ราคาต่อกก.", "ปริมาณกก.", "รายรับ", "รายจ่าย", "ค่าตัด",
    "เบิก", "สุทธิ", "วิธีรับ/จ่ายเงิน", "สถานะรายการ", "วันที่บันทึก", "หมายเหตุ", "แหล่งที่มา",
]
for col, header in enumerate(headers, 1):
    cell = ws.cell(17, col, header)
    cell.fill = PatternFill("solid", fgColor=green)
    cell.font = Font(bold=True, color=white)
    cell.alignment = Alignment(horizontal="center", vertical="center")

formulas = [
    '=TEXT(COUNTA(\'ฐานข้อมูล\'!$A:$A)-3+1,"TXN-0000")',
    "=C4", "=C5", "=C6", "=C7", '=IF(C8<>"",C8,"")', '=IF(C9<>"",C9,"")',
    "=C10", "=C11", "=F10", "=F11", "=F4", "=F5", "=F12", "=F6", "=F7", "=TODAY()", "=F8", '"เพิ่มข้อมูล"',
]
for col, formula in enumerate(formulas, 1):
    cell = ws.cell(18, col, formula)
    cell.fill = PatternFill("solid", fgColor=light_green)
    cell.border = Border(bottom=thin)
    cell.alignment = Alignment(horizontal="left", vertical="center")

for col in [2, 17]:
    ws.cell(18, col).number_format = "d/m/yyyy"
for col in [8, 9, 10, 11, 12, 13, 14]:
    ws.cell(18, col).number_format = "#,##0.00"

ws["A20"] = "วิธีใช้: กรอกข้อมูลด้านบนให้ครบจนสถานะเป็น “พร้อมบันทึก” แล้วคัดลอกทั้งแถว 18 ไปวางท้ายตารางในชีตฐานข้อมูล"
ws["A20"].font = Font(italic=True, color="60706A")
ws.merge_cells("A20:S20")

widths = {
    "A": 14, "B": 13, "C": 12, "D": 13, "E": 16, "F": 24, "G": 20,
    "H": 12, "I": 12, "J": 13, "K": 13, "L": 11, "M": 11, "N": 13,
    "O": 16, "P": 14, "Q": 13, "R": 28, "S": 16,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width

for row in [17, 18]:
    ws.row_dimensions[row].height = 24

for sheet in wb.worksheets:
    sheet.protection.sheet = False
    sheet.protection.enable = False

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(out)
print(out)
