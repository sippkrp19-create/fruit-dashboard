from __future__ import annotations

from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


SRC = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี.xlsx")
OUT = Path(r"C:\Users\sippk\OneDrive\เอกสาร\Open Design\dashboard\สวนผลไม้_บัญชี_ปรับประสิทธิภาพ.xlsx")

GREEN_DARK = "1F4E3D"
GREEN_LIGHT = "E9F4EC"
GOLD_LIGHT = "FFF3D6"
WHITE = "FFFFFF"
GRAY_TEXT = "60706A"

MAX_DATA_ROW = 2000


def ymd_season(date_value):
    if not date_value:
        return None
    year = getattr(date_value, "year", None)
    if not year:
        return None
    return year + 543 if year < 2400 else year


def set_header_style(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        if cell.value is not None:
            cell.fill = PatternFill("solid", fgColor=GREEN_DARK)
            cell.font = Font(bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center")


def money_formula(value_range, criteria: str):
    return f"SUMIFS({value_range},{criteria})"


def filtered_sum(value_col: str, extra_criteria: str = ""):
    base = (
        f"'ฐานข้อมูล'!${value_col}$5:${value_col}${MAX_DATA_ROW},"
        f"'ฐานข้อมูล'!$P$5:$P${MAX_DATA_ROW},\"ปกติ\","
        f"'ฐานข้อมูล'!$C$5:$C${MAX_DATA_ROW},$B$3"
    )
    if extra_criteria:
        base += f",{extra_criteria}"
    start_date = "DATE($B$3-543,$E$3,1)"
    month_criteria = (
        f",{extra_criteria}" if extra_criteria else ""
    ) + (
        f",'ฐานข้อมูล'!$B$5:$B${MAX_DATA_ROW},\">=\"&{start_date},"
        f"'ฐานข้อมูล'!$B$5:$B${MAX_DATA_ROW},\"<\"&EDATE({start_date},1)"
    )
    month_base = (
        f"'ฐานข้อมูล'!${value_col}$5:${value_col}${MAX_DATA_ROW},"
        f"'ฐานข้อมูล'!$P$5:$P${MAX_DATA_ROW},\"ปกติ\","
        f"'ฐานข้อมูล'!$C$5:$C${MAX_DATA_ROW},$B$3"
        f"{month_criteria}"
    )
    return f'=IF($E$3="ทั้งหมด",SUMIFS({base}),SUMIFS({month_base}))'


def build():
    wb = openpyxl.load_workbook(SRC)
    form = wb["เพิ่มข้อมูล"]
    db = wb["ฐานข้อมูล"]
    summary = wb["สรุป"]

    # Normalize season/year in database from the real date column.
    for row in range(5, db.max_row + 1):
        season = ymd_season(db.cell(row, 2).value)
        if season:
            db.cell(row, 3, season)
        db.cell(row, 10, f'=IF($E{row}="รายรับ",$H{row}*$I{row},0)')
        db.cell(row, 11, f'=IF($E{row}="รายจ่าย",$H{row}*$I{row},0)')
        db.cell(row, 14, f"=J{row}-K{row}-L{row}")

    # Add/refresh database table and filters.
    db.tables.clear()
    table_ref = f"A4:S{db.max_row}"
    table = Table(displayName="tblDatabase", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    db.add_table(table)
    db.auto_filter.ref = table_ref
    db.freeze_panes = "A5"
    db.sheet_view.showGridLines = False

    widths = {
        "A": 13, "B": 13, "C": 12, "D": 12, "E": 14, "F": 24, "G": 20,
        "H": 12, "I": 12, "J": 14, "K": 14, "L": 12, "M": 12, "N": 14,
        "O": 16, "P": 14, "Q": 14, "R": 34, "S": 22,
    }
    for col, width in widths.items():
        db.column_dimensions[col].width = width
    set_header_style(db, 4, 1, 19)
    for row in db.iter_rows(min_row=5, max_row=db.max_row):
        row[1].number_format = "d/m/yyyy"
        row[7].number_format = "#,##0.00"
        row[8].number_format = "#,##0.00"
        for idx in [9, 10, 11, 12, 13]:
            row[idx].number_format = "#,##0.00"
        row[16].number_format = "d/m/yyyy"

    # Rebuild Summary as the main read page.
    for row in summary.iter_rows():
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)
            cell.border = Border()
            cell.font = Font(name="Aptos", size=11)
            cell.alignment = Alignment(horizontal="left", vertical="center")
    summary.tables.clear()
    summary.data_validations.dataValidation = []
    summary.sheet_view.showGridLines = False
    summary.freeze_panes = "A5"
    summary["A1"] = "สวนผลไม้: สรุปบัญชี"
    summary["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=GREEN_DARK)
    summary["A2"] = "เลือกปี/เดือนเพื่อดูรายรับ รายจ่าย สุทธิ และรายการสำคัญจากฐานข้อมูล"
    summary["A2"].font = Font(size=10, color=GRAY_TEXT)
    summary["A3"] = "ปี"
    summary["B3"] = 2569
    summary["C3"] = "เดือน"
    summary["D3"] = "ทั้งหมด"
    summary["E3"] = '=IF(D3="ทั้งหมด","ทั้งหมด",MATCH(D3,$Y$2:$Y$13,0))'
    summary.column_dimensions["E"].hidden = True
    for cell in ["A3", "C3"]:
        summary[cell].font = Font(bold=True, color=GREEN_DARK)
    for cell in ["B3", "D3"]:
        summary[cell].fill = PatternFill("solid", fgColor=GOLD_LIGHT)
        summary[cell].font = Font(bold=True)
        summary[cell].alignment = Alignment(horizontal="center")

    years = sorted({db.cell(row, 3).value for row in range(5, db.max_row + 1) if isinstance(db.cell(row, 3).value, int)})
    if 2569 not in years:
        years.insert(0, 2569)
    months = ["ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.", "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
    for idx, year in enumerate(years[:10], 1):
        summary[f"X{idx}"] = year
    summary["Y1"] = "ทั้งหมด"
    for idx, month in enumerate(months, 2):
        summary[f"Y{idx}"] = month
    for col in ["X", "Y"]:
        summary.column_dimensions[col].hidden = True
    dv_year = DataValidation(type="list", formula1=f"=$X$1:$X${max(len(years[:10]), 1)}", allow_blank=False)
    dv_month = DataValidation(type="list", formula1="=$Y$1:$Y$13", allow_blank=False)
    summary.add_data_validation(dv_year)
    summary.add_data_validation(dv_month)
    dv_year.add(summary["B3"])
    dv_month.add(summary["D3"])

    # KPI block.
    kpis = [
        ("รายรับรวม", filtered_sum("J")),
        ("รายจ่ายรวม", f"={filtered_sum('K')[1:]}+{filtered_sum('L')[1:]}"),
        ("กำไรสุทธิ", "=B5-B6"),
        ("ปริมาณกก.รวม", filtered_sum("I")),
        ("ราคาเฉลี่ย/กก.", "=IFERROR(B5/B8,0)"),
        ("ค่าตัดรวม", filtered_sum("L")),
        ("เบิกรวม", filtered_sum("M")),
    ]
    summary["A4"] = "KPI"
    summary["B4"] = "ค่า"
    for idx, (label, formula) in enumerate(kpis, 5):
        summary.cell(idx, 1, label)
        summary.cell(idx, 2, formula)

    # Product summary.
    products = ["ทุเรียน", "มังคุด", "ลองกอง", "ยังไม่ชัว"]
    for col, header in enumerate(["สินค้า", "รายรับ", "รายจ่าย", "สุทธิ"], 4):
        summary.cell(4, col, header)
    for idx, product in enumerate(products, 5):
        summary.cell(idx, 4, product)
        product_criteria = f"'ฐานข้อมูล'!$D$5:$D${MAX_DATA_ROW},D{idx}"
        summary.cell(idx, 5, filtered_sum("J", product_criteria))
        summary.cell(idx, 6, f"={filtered_sum('K', product_criteria)[1:]}+{filtered_sum('L', product_criteria)[1:]}")
        summary.cell(idx, 7, f"=E{idx}-F{idx}")

    # Expense names sorted by current overall amount.
    expense_totals = defaultdict(float)
    for row in range(5, db.max_row + 1):
        name = db.cell(row, 7).value
        if name:
            price = db.cell(row, 8).value or 0
            qty = db.cell(row, 9).value or 0
            if db.cell(row, 5).value == "รายจ่าย":
                expense_totals[str(name)] += float(price) * float(qty)
    top_names = [name for name, _ in sorted(expense_totals.items(), key=lambda item: item[1], reverse=True)[:12]]
    summary["I4"] = "ชื่อ/รายการ"
    summary["J4"] = "รายจ่าย"
    for idx, name in enumerate(top_names, 5):
        summary.cell(idx, 9, name)
        name_criteria = f"'ฐานข้อมูล'!$G$5:$G${MAX_DATA_ROW},I{idx}"
        summary.cell(idx, 10, filtered_sum("K", name_criteria))

    # Monthly summary.
    summary["A14"] = "สรุปรายเดือน"
    summary["A14"].font = Font(bold=True, color=GREEN_DARK, size=14)
    for col, header in enumerate(["เดือน", "รายรับ", "รายจ่าย", "สุทธิ"], 1):
        summary.cell(15, col, header)
    for idx, month_num in enumerate(range(1, 13), 16):
        summary.cell(idx, 1, f'=TEXT(DATE($B$3-543,{month_num},1),"mmm yyyy")')
        start_date = f"DATE($B$3-543,{month_num},1)"
        next_month = f"EDATE({start_date},1)"
        summary.cell(idx, 2, f'=SUMIFS(\'ฐานข้อมูล\'!$J$5:$J${MAX_DATA_ROW},\'ฐานข้อมูล\'!$P$5:$P${MAX_DATA_ROW},"ปกติ",\'ฐานข้อมูล\'!$C$5:$C${MAX_DATA_ROW},$B$3,\'ฐานข้อมูล\'!$B$5:$B${MAX_DATA_ROW},">="&{start_date},\'ฐานข้อมูล\'!$B$5:$B${MAX_DATA_ROW},"<"&{next_month})')
        summary.cell(idx, 3, f'=SUMIFS(\'ฐานข้อมูล\'!$K$5:$K${MAX_DATA_ROW},\'ฐานข้อมูล\'!$P$5:$P${MAX_DATA_ROW},"ปกติ",\'ฐานข้อมูล\'!$C$5:$C${MAX_DATA_ROW},$B$3,\'ฐานข้อมูล\'!$B$5:$B${MAX_DATA_ROW},">="&{start_date},\'ฐานข้อมูล\'!$B$5:$B${MAX_DATA_ROW},"<"&{next_month})+SUMIFS(\'ฐานข้อมูล\'!$L$5:$L${MAX_DATA_ROW},\'ฐานข้อมูล\'!$P$5:$P${MAX_DATA_ROW},"ปกติ",\'ฐานข้อมูล\'!$C$5:$C${MAX_DATA_ROW},$B$3,\'ฐานข้อมูล\'!$B$5:$B${MAX_DATA_ROW},">="&{start_date},\'ฐานข้อมูล\'!$B$5:$B${MAX_DATA_ROW},"<"&{next_month})')
        summary.cell(idx, 4, f"=B{idx}-C{idx}")

    # Styling and tables.
    for row in [4, 15]:
        set_header_style(summary, row, 1, 10)
    for row in summary.iter_rows(min_row=5, max_row=27, min_col=1, max_col=10):
        for cell in row:
            if cell.column in [2, 3, 4, 5, 6, 7, 10]:
                cell.number_format = "#,##0.00"
            if cell.row % 2 == 0 and cell.value is not None:
                cell.fill = PatternFill("solid", fgColor=GREEN_LIGHT)
    for col, width in {
        "A": 20, "B": 16, "C": 5, "D": 16, "E": 16, "F": 16, "G": 16,
        "H": 5, "I": 22, "J": 16,
    }.items():
        summary.column_dimensions[col].width = width
    for ref, name in [("A4:B11", "tblKPI"), ("D4:G8", "tblByProduct"), (f"I4:J{4+len(top_names)}", "tblExpenseByName"), ("A15:D27", "tblByMonth")]:
        table = Table(displayName=name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True, showColumnStripes=False)
        summary.add_table(table)

    # Improve form usability and keep editable.
    form.sheet_view.showGridLines = False
    form["A2"] = "กรอกช่องสีเหลือง แล้วคัดลอกแถวตัวอย่างไปเพิ่มท้ายตารางฐานข้อมูล"
    form["A2"].font = Font(color=GRAY_TEXT, italic=True)
    for sheet in wb.worksheets:
        sheet.protection.sheet = False
        sheet.protection.enable = False

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.calculation.calcMode = "auto"
    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    print(build())
